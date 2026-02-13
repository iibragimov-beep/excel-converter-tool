import re
import openpyxl
from openpyxl.utils import get_column_letter
import streamlit as st
import io


def format_button(button):
    # ... (Your original logic remains exactly the same) ...
    key_match = re.search(r'key:(\d+)', button)
    if not key_match:
        return None, None
    key = int(key_match.group(1))

    value_match = re.search(r'value:([^,|]+)', button)
    if not value_match:
        return None, None
    button_type = value_match.group(1)

    label_match = re.search(r'LabelName=([^;,\|]*)', button)
    label = label_match.group(1) if label_match else ""

    output = ""

    if button_type == 'vu-display':
        extension_match = re.search(r'Extension=([^;,\|]*)', button)
        if extension_match:
            output = f"vu-display={extension_match.group(1)}"
    elif button_type == 'autodial':
        dialed_match = re.search(r'DialedNumber=([^;,\|]*)', button)
        if dialed_match:
            output = f"autodial={dialed_match.group(1)}"
    elif button_type == 'q-calls':
        group_match = re.search(r'EmployeeGroup=(\d+)', button)
        if group_match:
            output = f"q-calls={group_match.group(1)}"
    elif button_type == 'busy-ind':
        bi_match = re.search(r'BIExtension=(\d+)', button)
        if bi_match:
            output = f"busy-ind={bi_match.group(1)}"
    elif button_type == 'aut-msg-wt':
        mwi_match = re.search(r'MWILampExtension=(\d+)', button)
        if mwi_match:
            output = f"aut-msg-wt={mwi_match.group(1)}"
    elif button_type == 'brdg-appr':
        ext_match = re.search(r'Ext=(\d+)', button)
        button_match = re.search(r'Button=([^;,\|]+)', button)
        if ext_match and button_match:
            output = f"brdg-appr={ext_match.group(1)},{button_match.group(1)}"
    elif button_type == 'sip-sobsrv':
        listen_match = re.search(r'ListenOnly=(true|false)', button)
        coach_match = re.search(r'Coach=(true|false)', button)
        components = []
        if listen_match and listen_match.group(1) == 'true':
            components.append("ListenOnly")
        if coach_match and coach_match.group(1) == 'true':
            components.append("Coach")
        if components:
            output = f"sip-sobsrv {' '.join(components)}"
    else:
        output = button_type

    if label:
        output = f"LabelName={label} {output}"

    return key, output


def get_max_keys(ws_input):
    max_key = 0
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        button_features = row[6]
        if button_features:
            for button in button_features.split('|'):
                key_match = re.search(r'key:(\d+)', button)
                if key_match:
                    max_key = max(max_key, int(key_match.group(1)))
    return max_key


def process_input_excel(input_file, output_file):
    wb_input = openpyxl.load_workbook(input_file)
    ws_input = wb_input['Endpoints']

    max_keys = get_max_keys(ws_input) or 52

    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = "Avaya Buttons"

    headers = (
            ['Name', 'Number', 'PermissionSet'] +
            [f'Key {i}' for i in range(1, max_keys + 1)] +
            ['Profile', 'GroupId', 'BridgedCallAlerting', 'DialingOption',
             'HeadsetSignaling', 'ButtonClicks', 'PhoneScreen', 'Redial',
             'AudioPath', 'UserPrefferedLanguage']
    )

    for col, header in enumerate(headers, start=1):
        ws_output.cell(row=1, column=col, value=header)

    for row_idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=True), start=2):
        name, number, permission_set = row[0], row[1], row[5]
        button_features, profile, group_id = row[6], row[7], row[13]
        bridged_call_alerting, dialing_option = row[15], row[18]
        headset_signaling, audio_path, button_clicks = row[19], row[20], row[21]
        phone_screen, user_pref_lang, redial = row[22], row[26], row[28]

        ws_output[f"A{row_idx}"] = name
        ws_output[f"B{row_idx}"] = number
        ws_output[f"C{row_idx}"] = permission_set
        ws_output[f"{get_column_letter(max_keys + 4)}{row_idx}"] = profile
        ws_output[f"{get_column_letter(max_keys + 5)}{row_idx}"] = group_id
        ws_output[f"{get_column_letter(max_keys + 6)}{row_idx}"] = bridged_call_alerting
        ws_output[f"{get_column_letter(max_keys + 7)}{row_idx}"] = dialing_option
        ws_output[f"{get_column_letter(max_keys + 8)}{row_idx}"] = headset_signaling
        ws_output[f"{get_column_letter(max_keys + 9)}{row_idx}"] = button_clicks
        ws_output[f"{get_column_letter(max_keys + 10)}{row_idx}"] = phone_screen
        ws_output[f"{get_column_letter(max_keys + 11)}{row_idx}"] = redial
        ws_output[f"{get_column_letter(max_keys + 12)}{row_idx}"] = audio_path
        ws_output[f"{get_column_letter(max_keys + 13)}{row_idx}"] = user_pref_lang

        if button_features:
            for button in button_features.split('|'):
                key, output = format_button(button)
                if key and output is not None and key <= max_keys:
                    col = get_column_letter(key + 3)
                    ws_output[f"{col}{row_idx}"] = output

    # CHANGED: This now saves to the output_file object (the buffer)
    wb_output.save(output_file)


# --- STREAMLIT INTERFACE ---
st.set_page_config(page_title="Avaya Button Converter")
st.title("📞 Avaya Button Processor")
st.write("Upload your 'Endpoints' Excel file to generate the formatted button map.")

uploaded_file = st.file_uploader("Choose the input Excel file", type="xlsx")

if uploaded_file:
    try:
        # Create a buffer for the output file
        output_buffer = io.BytesIO()

        # Process the file
        with st.spinner("Processing buttons..."):
            process_input_excel(uploaded_file, output_buffer)

        st.success("Conversion complete!")

        # Provide download button
        st.download_button(
            label="💾 Download Processed Output",
            data=output_buffer.getvalue(),
            file_name="Avaya_Buttons_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"An error occurred: {e}")
        st.info("Check if your input file has a sheet named 'Endpoints'.")
