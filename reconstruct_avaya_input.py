import openpyxl
import re


def reverse_format(simplified, key, display_format=9):
    simplified = (simplified or "").strip()
    if not simplified:
        return None
    # =========================================================
    # LOGOUT-OVR — MUST BE DETECTED BEFORE LABEL PARSING
    # =========================================================
    if "logout-ovr" in simplified:
        label = ""
        if simplified.startswith("LabelName="):
            label = simplified[len("LabelName="):].replace("logout-ovr", "").strip()
        return (
            f"params:LabelName={label},"
            f"key:{key},value:logout-ovr"
        )

    label_value = ""
    rest = simplified

    # -------------------------------
    # Extract LabelName if present
    # -------------------------------
    if simplified.startswith("LabelName="):
        known_types = (
            r'\b(vu-display|autodial|q-calls|busy-ind|aut-msg-wt|brdg-appr|sip-sobsrv|'
            r'call-appr|call-fwd|send-calls|agnt-login|auto-in|aux-work|after-call|'
            r'mct-act|call-park|call-unpk|call-pkup)\b'
        )
        match = re.search(known_types, simplified)
        if match:
            label_end = match.start()
            label_value = simplified[len("LabelName="):label_end].rstrip('; ').strip()
            rest = simplified[label_end:].strip()
        else:
            label_value = simplified[len("LabelName="):].rstrip('; ').strip()
            rest = ""

    # =========================================================
    # SIP-SOBSRV — MUST BE DETECTED FIRST (TRAILING TOKENS)
    # =========================================================
    if "sip-sobsrv" in rest:
        label = label_value or ""
        return (
            f"params:ListenOnly=true;"
            f"Coach=false;"
            f"LabelName={label},"
            f"key:{key},value:sip-sobsrv"
        )

    # -------------------------------
    # Parse remaining buttons
    # -------------------------------
    button_type = ""
    param = ""
    if '=' in rest:
        button_type, param = rest.split('=', 1)
        param = param.strip()
    else:
        button_type = rest.strip()

    # =========================================================
    # ISOLATED BUTTONS (EARLY RETURN — NO SHARED LOGIC)
    # =========================================================

    # busy-ind
    if button_type == "busy-ind" and param:
        label = label_value or ""
        return f"params:LabelName={label};BIExtension={param},key:{key},value:busy-ind"

    # vu-display
    if button_type == "vu-display" and param:
        label = label_value or ""
        return (
            f"params:LabelName={label};"
            f"DisplayFormat={display_format};Extension={param},"
            f"key:{key},value:vu-display"
        )

    # autodial — Overhead Paging
    if button_type == "autodial" and param and label_value == "Overhead Paging":
        return (
            f"params:DialedNumber={param};"
            f"LabelName={label_value},"
            f"key:{key},value:autodial"
        )

    # autodial — generic
    if button_type == "autodial" and param:
        label = label_value or ""
        return (
            f"params:LabelName={label};"
            f"DialedNumber={param},"
            f"key:{key},value:autodial"
        )

    # aut-msg-wt
    if button_type == "aut-msg-wt" and param:
        label = label_value or ""
        return (
            f"params:LabelName={label};"
            f"MWILampExtension={param},"
            f"key:{key},value:aut-msg-wt"
        )

    # brdg-appr
    if button_type == "brdg-appr" and param:
        label = label_value or ""
        parts = param.split(",")
        if len(parts) == 2:
            ext, button = parts
            return (
                f"params:LabelName={label};"
                f"Button={button};Ext={ext},"
                f"key:{key},value:brdg-appr"
            )

    # q-calls
    if button_type == "q-calls" and param:
        label = label_value or ""
        return (
            f"params:LabelName={label};"
            f"EmployeeGroup={param},"
            f"key:{key},value:q-calls"
        )

    # =========================================================
    # FALLBACK / LEGACY BUTTONS (UNCHANGED)
    # =========================================================

    params_body = ""
    separator = ","

    if button_type == "call-pkup":
        params_body = "Tone=half-ring;RingerType=no-ring;"

    label_part = (
        f"LabelName={label_value}{separator}"
        if label_value else f"LabelName={separator}"
    )

    prefix = "params:"

    if button_type == "call-pkup":
        full = f"{prefix}{params_body}{label_part}key:{key},value:{button_type}"
    else:
        full = f"{prefix}{label_part}{params_body}key:{key},value:{button_type}"

    full = re.sub(r';+', ';', full)
    full = re.sub(r';key:', 'key:', full)
    return full.rstrip(';')


def reconstruct_input(output_file, reconstructed_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Avaya Buttons']

    headers = [cell.value for cell in ws[1] if cell.value]
    key_headers = [h for h in headers if h.startswith("Key ")]
    max_keys = max(int(h.split()[1]) for h in key_headers)

    original_headers = [
        "Name", "Number", "AgentPhone", "Password", "AutoAnswer", "PermissionSet",
        "ButtonFeatures", "Profile", "VmNumber", "Room", "Floor", "Building",
        "CoveragePathId", "GroupId", "MultiRegistration", "BridgedCallAlerting",
        "BlockEnhancedCallPickupAlerting", "PhoneScreenOnCalling", "DialingOption",
        "HeadsetSignaling", "AudioPath", "ButtonClicks", "PhoneScreen",
        "BackgroundLogo", "PersonalizedRinging", "InternalCallCdr",
        "UserPrefferedLanguage", "TimeFormat", "Redial"
    ]

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Endpoints"

    for col, h in enumerate(original_headers, start=1):
        ws_new.cell(row=1, column=col, value=h)

    header_map = {h: i + 1 for i, h in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        row = [""] * len(original_headers)

        field_map = {
            "Name": 0,
            "Number": 1,
            "PermissionSet": 5,
            "Profile": 7,
            "GroupId": 13,
            "BridgedCallAlerting": 15,
            "DialingOption": 18,
            "HeadsetSignaling": 19,
            "ButtonClicks": 21,
            "PhoneScreen": 22,
            "Redial": 28,
        }

        for field, idx in field_map.items():
            if field in header_map:
                row[idx] = ws.cell(row=row_idx, column=header_map[field]).value

        button_parts = []
        vu_display_count = 0

        for k in range(1, max_keys + 1):
            col = header_map.get(f"Key {k}")
            if not col:
                continue

            simplified = ws.cell(row=row_idx, column=col).value
            if not simplified:
                continue

            if "vu-display" in simplified:
                vu_display_count += 1
                display_format = 8 + vu_display_count
                part = reverse_format(simplified, k, display_format)
            else:
                part = reverse_format(simplified, k)

            if part:
                button_parts.append(part)

        row[6] = "|".join(button_parts)

        # Required constants
        row[2]  = "Local"
        row[3]  = "24688"
        row[8]  = "59030"
        row[12] = "IXM"
        row[14] = "1"
        row[16] = "TRUE"
        row[17] = "false"
        row[24] = "CLASSIC_TONE_3"
        row[25] = "USE_PROFILE"
        row[27] = "HOUR_12"

        for col, val in enumerate(row, start=1):
            ws_new.cell(row=row_idx, column=col, value=val)

    wb_new.save(reconstructed_file)
    print(f"Reconstructed file saved as {reconstructed_file}")


if __name__ == "__main__":
    reconstruct_input("Output.xlsx", "Reconstructed Test Input.xlsx")
